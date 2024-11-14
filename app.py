import streamlit as st
import pandas as pd
import re
import os
from PIL import Image
import pytesseract
import numpy as np
import openpyxl
from datetime import datetime


# -------------- SETTINGS --------------

page_title = "Automatizated Petty cash"
page_icon = "💳"  # emojis: https://www.webfx.com/tools/emoji-cheat-sheet/
layout = "wide"
euro_symbol = '\u20AC'
total_expenses = 0
final_price = 0
df_expense = ""
css = "style/main.css"
url_logo = "assets/Imagen1.png"
# Ruta a la carpeta de documentos
folder_path = 'invoices'

st.set_page_config(
    page_title="Automatizated Petty cash",
    page_icon=page_icon,
    layout=layout,
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://valerapp.com/contact/',
        'Report a bug': "https://valerapp.com/contact/",
        'About': "# This is an *extremely* cool app!"
    }
)

hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

# Funciones


def get_files_in_folder(folder_path):
    files = []
    # Itera sobre todos los archivos y subdirectorios en la carpeta
    for root, dirs, filenames in os.walk(folder_path):
        for filename in filenames:
            # Añadir la ruta completa del archivo a la lista de archivos
            files.append(os.path.join(root, filename))
    return files

def extraer_texto_de_imagenes(file_path):
    try:
        # Abre la imagen y extrae el texto
        text_read = pytesseract.image_to_string(Image.open(file_path))
        print(f'Texto extraído de {os.path.basename(file_path)}:')
        print(text_read)
        print('-' * 50)
        # Aplicar las funciones para extraer la información
        proveedor, fecha, factura, monto = extraer_texto(text_read)
        return proveedor, fecha, factura, monto
    except Exception as e:
        print(f'Error procesando el archivo {file_path}: {e}')
        return None, None, None, None


def extraer_proveedor(text_read):        
    # Patrón para extraer el concepto/proveedor para pago móvil
    concepto_pattern = r'CONCEPTO\s*([^\n]+)'
    concepto_match = re.search(concepto_pattern, text_read)
    proveedor = concepto_match.group(1) if concepto_match else None
    
    # Patrón para extraer el nombre del lugar para facturas
    lugar_pattern = r'IDENTIFICACION RECEPTOR\s[^\n]+\n([^\n]+)'
    lugar_match = re.search(lugar_pattern, text_read)
    nombre_lugar = lugar_match.group(1) if lugar_match else None

    # Si es una factura, usar el nombre del lugar. Si es un pago móvil, usar el concepto o 'Pago móvil'.
    nombre_lugar_proveedor = nombre_lugar if nombre_lugar else proveedor
    return nombre_lugar_proveedor

def extraer_fecha(text_read):
    # Patrón para extraer la fecha
    fecha_pattern = r'FECHA\s*\n\s*(\d{2}/\d{2}/\d{4})'
    fecha_match = re.search(fecha_pattern, text_read)
    fecha = fecha_match.group(1) if fecha_match else None

    # Convertir la fecha al formato dd/mm/yyyy (si es necesario)
    if fecha:
        try:
            fecha = datetime.strptime(fecha, '%d/%m/%Y').strftime('%d/%m/%Y')
        except ValueError:
            pass
    return fecha


def extraer_factura(text_read):
    # Patrón para extraer el número de referencia del pago móvil
    factura_pattern = r'NUMERO DE REFERENCIA\s+[^\n]*\n(\d+)'
    factura_match = re.search(factura_pattern, text_read)
    factura = factura_match.group(1) if factura_match else None
    return factura

def extraer_monto(text_read):
    # Patrón para extraer el monto en bolívares
    monto_pattern = r'MONTO DE LA OPERACION\s*[^\n]*\nBs\.?\s*([\d\.]+,\d{2})'
    monto_matches = re.findall(monto_pattern, text_read)
    monto = monto_matches[-1].replace('.', '').replace(',', '.') if monto_matches else None  # Obtener el último monto encontrado
    return monto    

def extraer_texto(text_read):
    proveedor = extraer_proveedor(text_read)
    fecha = extraer_fecha(text_read)
    factura = extraer_factura(text_read)
    monto = extraer_monto(text_read)
    
    # Verificar que todos los datos hayan sido extraídos correctamente
    if None in (proveedor, fecha, factura, monto):
        raise ValueError("Datos extraídos son incompletos")
    
    return proveedor, fecha, factura, monto



# -------------- Frontend code ----------------

# Crear las carpetas necesarias si no existen
os.makedirs('invoices', exist_ok=True)
os.makedirs('processed_invoices', exist_ok=True)
os.makedirs('caja_chica', exist_ok=True)

# title
st.title("Automatizated Petty cash")

# Sección de información de la factura
with st.container():
    cc1, cc2 = st.columns(2)
    cc1.image(url_logo, width=100)
    from_who = cc1.text_input("Responsable:", placeholder="Nombre del responsable")
    cc2.subheader("Datos del responsable")
    num_invoice = cc2.text_input("#", placeholder="Cédula del responsable", max_chars=8)
    date_invoice = cc2.date_input("Fecha:")
    due_date = cc2.number_input("Monto $:", min_value=0.0, format="%.2f")

# Botón para subir fotos
uploaded_files = st.file_uploader("Subir fotos", type=["png", "jpg", "jpeg"], accept_multiple_files=True)

# Mostrar archivos subidos
if uploaded_files:
    st.write("Archivos subidos:")
    for uploaded_file in uploaded_files:
        # Guardar archivo en la carpeta invoices
        with open(os.path.join("invoices", uploaded_file.name), "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.success(f"Archivo guardado: {uploaded_file.name}")

# Validar que todos los campos requeridos estén llenos
if st.button("Crear archivo Excel"):
    if not from_who:
        st.error("El campo 'Responsable' es requerido.")
    elif not num_invoice:
        st.error("El campo 'Cédula del responsable' es requerido.")
    elif not due_date:
        st.error("El campo 'Monto $' es requerido.")
    else:
        st.success(f"Hola, {from_who}!")
        folder_path = 'invoices'
        files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

        # Crear una lista para almacenar los datos extraídos
        all_data = []

        for file in files:
            try:
                proveedor, fecha, factura, monto = extraer_texto_de_imagenes(file)
                if None not in (proveedor, fecha, factura, monto):
                    # Asegurarse de que el monto se trata como un número
                    if monto is not None:
                        monto = float(monto)

                    # Añadir los datos extraídos a la lista
                    all_data.append([proveedor, fecha, factura, monto])

                    # Mover el archivo de imagen procesado a la carpeta processed_invoices
                    new_file_path = os.path.join('processed_invoices', os.path.basename(file))
                    os.rename(file, new_file_path)
                else:
                    st.error(f"Error procesando el archivo {file}: Datos extraídos son incompletos.")
            except Exception as e:
                st.error(f"Error procesando el archivo {file}: {e}")

        # Abrir el archivo Excel base
        base_excel_path = 'base/caja_chica_base.xlsx'
        if not os.path.exists(base_excel_path):
            st.error(f"El archivo base {base_excel_path} no existe.")
        else:
            workbook = openpyxl.load_workbook(base_excel_path)
            worksheet = workbook.active

            # Agregar la información del responsable, cédula, fecha actual y monto otorgado
            worksheet['C9'] = from_who  # Nombre del responsable
            worksheet['K9'] = num_invoice  # Cédula del responsable
            worksheet['K10'] = datetime.now().strftime('%d/%m/%Y')  # Fecha actual
            worksheet['K11'] = due_date  # Monto otorgado en $

            # Comenzar a agregar datos desde la celda C14
            start_row = 14
            start_column = 3  # Columna C

            # Agregar datos extraídos al archivo Excel base, aplicando formato de moneda a la columna F
            for row_index, data in enumerate(all_data, start=start_row):
                col_index = start_column
                for value in data:
                    cell = worksheet.cell(row=row_index, column=col_index, value=value)
                    if col_index == 6 and value is not None:  # Columna F y si el valor no es None
                        cell.number_format = '[$Bs S-VE] #,##0.00'  # Formato de moneda
                    col_index += 1
                    if col_index == 6:  # Ignorar la columna F
                        col_index += 1

            # Aplicar el formato de moneda en dólares a la columna K
            for row in range(start_row, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=11)  # Columna K es la columna 11
                cell.number_format = '[$$-409]#.##0,00'  # Formato de moneda en dólares

            # Asegurarnos de que la celda K11 tenga el formato correcto
            worksheet['K11'].number_format = '[$$-409]#.##0,00'

            # Obtener la fecha actual para usarla en el nombre del archivo
            fecha_actual = datetime.now().strftime('%d-%m-%Y')

            # Guardar los cambios en un nuevo archivo Excel con el nombre incluyendo la fecha
            updated_excel_path = f'caja_chica/caja_chica_{fecha_actual}.xlsx'
            workbook.save(updated_excel_path)
            st.success(f"Datos agregados y guardados en: {updated_excel_path}")

            # Botón para descargar el archivo Excel
            with open(updated_excel_path, 'rb') as file:
                btn = st.download_button(
                    label="Descargar caja chica",
                    data=file,
                    file_name=updated_excel_path,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

st.warning("Cambiar el precio del dólar al del día en la casilla de la suma total en dólares")
